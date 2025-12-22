import requests
from requests.auth import HTTPBasicAuth
import json
import getpass
import sys
import signal
import argparse
import xml.etree.ElementTree as ET
import html
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# Global flag for graceful shutdown
interrupted = False

def get_progress_filename(platform, username):
    """Generate progress filename based on platform and username"""
    safe_username = "".join(c for c in username if c.isalnum() or c in ('-', '_')).lower()
    return f"tag_report_progress_{platform}_{safe_username}.json"

def save_progress(processed_tags, platform, username, silent=False):
    """Save current progress to a JSON file"""
    progress_file = get_progress_filename(platform, username)

    progress_data = {
        "platform": platform,
        "username": username,
        "processed_tags": processed_tags,
        "timestamp": datetime.now().isoformat()
    }

    try:
        with open(progress_file, 'w') as f:
            json.dump(progress_data, f)
        if not silent:
            print(f"\nProgress saved! ({len(processed_tags)} tags processed)")
    except (OSError, IOError) as e:
        print(f"\nWARNING: Failed to save progress: {e}")

def load_progress(platform, username):
    """Load progress from JSON file if it exists"""
    progress_file = get_progress_filename(platform, username)
    if os.path.exists(progress_file):
        try:
            with open(progress_file, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError, OSError, ValueError) as e:
            print(f"WARNING: Failed to load progress file: {e}")
            return None
    return None

def delete_progress(platform, username):
    """Delete the progress file"""
    progress_file = get_progress_filename(platform, username)
    if os.path.exists(progress_file):
        os.remove(progress_file)

def signal_handler(_sig, _frame):
    """Handle Ctrl+C gracefully"""
    global interrupted
    interrupted = True
    print("\n\nInterrupt received! Saving progress...")

# Register the signal handler for Ctrl+C
signal.signal(signal.SIGINT, signal_handler)

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Qualys Tag Report Generator')
parser.add_argument('--platform', type=str, help='Platform (US1, US2, US3, US4, UK, EU1, EU2, EU3, IN, CA, AE, AU, KSA)')
parser.add_argument('--username', type=str, help='Qualys username')
parser.add_argument('--password', type=str, help='Qualys password')
args = parser.parse_args()

# Ask for the platform selection and username first (if not provided via args)
if args.platform:
    platform = args.platform.upper()
else:
    print("\nOptions: US1, US2, US3, US4, UK, EU1, EU2, EU3, IN, CA, AE, AU, KSA\n")
    platform = input("What platform is your account on? ").upper()

# Define gateway URLs for each platform
gateway_urls = {
    "US1": "https://gateway.qg1.apps.qualys.com",
    "US2": "https://gateway.qg2.apps.qualys.com",
    "US3": "https://gateway.qg3.apps.qualys.com",
    "US4": "https://gateway.qg4.apps.qualys.com",
    "UK": "https://gateway.qg1.apps.qualys.co.uk",
    "EU1": "https://gateway.qg1.apps.qualys.eu",
    "EU2": "https://gateway.qg2.apps.qualys.eu",
    "EU3": "https://gateway.qg3.apps.qualys.it",
    "IN": "https://gateway.qg1.apps.qualys.in",
    "CA": "https://gateway.qg1.apps.qualys.ca",
    "AE": "https://gateway.qg1.apps.qualys.ae",
    "AU": "https://gateway.qg1.apps.qualys.com.au",
    "KSA": "https://gateway.qg1.apps.qualysksa.com"
}

# Define Qualys API URLs for QPS (different from gateway)
qualys_api_urls = {
    "US1": "https://qualysapi.qualys.com",
    "US2": "https://qualysapi.qg2.apps.qualys.com",
    "US3": "https://qualysapi.qg3.apps.qualys.com",
    "US4": "https://qualysapi.qg4.apps.qualys.com",
    "UK": "https://qualysapi.qg1.apps.qualys.co.uk",
    "EU1": "https://qualysapi.qualys.eu",
    "EU2": "https://qualysapi.qg2.apps.qualys.eu",
    "EU3": "https://qualysapi.qg3.apps.qualys.it",
    "IN": "https://qualysapi.qg1.apps.qualys.in",
    "CA": "https://qualysapi.qg1.apps.qualys.ca",
    "AE": "https://qualysapi.qg1.apps.qualys.ae",
    "AU": "https://qualysapi.qg1.apps.qualys.com.au",
    "KSA": "https://qualysapi.qg1.apps.qualysksa.com"
}

# Select the correct gateway URL
if platform in gateway_urls:
    gateway_url = gateway_urls[platform]
    qualys_api_url = qualys_api_urls[platform]
else:
    print("Invalid platform selection. Exiting")
    exit(1)

if args.username:
    username = args.username
else:
    username = input("Enter your username: ")

# Check for existing progress
existing_progress = load_progress(platform, username)
resume_from_progress = False

if existing_progress:
    # Check if progress is older than 24 hours
    try:
        progress_timestamp = datetime.fromisoformat(existing_progress.get('timestamp'))
        age_hours = (datetime.now() - progress_timestamp).total_seconds() / 3600

        if age_hours > 24:
            print("\n" + "="*70)
            print("STALE PROGRESS DETECTED")
            print("="*70)
            print(f"Progress file is {age_hours:.1f} hours old (saved: {existing_progress.get('timestamp')})")
            print("Progress older than 24 hours is considered stale and will be discarded.")
            print("Starting fresh session...")
            print("="*70)
            delete_progress(platform, username)
        else:
            print("\n" + "="*70)
            print("PREVIOUS SESSION DETECTED")
            print("="*70)
            print(f"Platform: {platform}")
            print(f"Username: {username}")
            print(f"Tags processed: {len(existing_progress.get('processed_tags', []))}")
            print(f"Last saved: {existing_progress.get('timestamp')} ({age_hours:.1f} hours ago)")
            print("="*70)

            resume_choice = input("\nDo you want to resume from saved progress? (yes/no): ").strip().lower()
            if resume_choice in ['yes', 'y']:
                resume_from_progress = True
                print(f"\nResuming session for {username} on {platform}...")
            else:
                print("\nStarting fresh session (deleting saved progress)...")
                delete_progress(platform, username)
    except (ValueError, TypeError):
        print("\n" + "="*70)
        print("INVALID PROGRESS DETECTED")
        print("="*70)
        print("Progress file has invalid timestamp. Starting fresh session...")
        print("="*70)
        delete_progress(platform, username)

# Get password
if args.password:
    password = args.password
elif not resume_from_progress:
    password = getpass.getpass("Enter your password: ")
else:
    password = getpass.getpass(f"Enter password for {username}: ")

# Define the authentication URL using the gateway URL
auth_url = f"{gateway_url}/auth"

# Authentication headers and data for JWT
auth_headers = {
    "Content-Type": "application/x-www-form-urlencoded"
}
auth_data = {
    "username": username,
    "password": password,
    "token": "true"
}

# Perform authentication to get JWT
try:
    auth_response = requests.post(auth_url, headers=auth_headers, data=auth_data, timeout=60)
except requests.exceptions.Timeout:
    print("ERROR: Authentication request timed out after 60 seconds.")
    print("Please check your network connection and try again.")
    exit(1)
except requests.exceptions.RequestException as e:
    print(f"ERROR: Network error during authentication: {e}")
    exit(1)

if auth_response.status_code != 201:
    print(f"Authentication failed. (HTTP {auth_response.status_code}).")
    print("Response from server:")
    print(auth_response.text or "No additional error details provided.")
    exit(1)

# Extract JWT token from response body (plain string)
jwt_token = auth_response.text.strip()
print("\nAuthentication successful.")

# ============================================================================
# Fetch all tags using the search endpoint
# ============================================================================

# Define the tag search URL (tags use /qps/rest/2.0/ path with Basic Auth)
# Try using the Qualys API URL instead of gateway URL
tag_url = f"{qualys_api_url}/qps/rest/2.0/search/am/tag"

tag_headers = {
    "Content-Type": "text/xml"
}

print("\nFetching tags...")

all_tags = []
page_number = 0
page_size = 100  # Default page size for tags

while True:
    # Prepare the request body as XML with pagination preferences
    # startFromOffset must be >= 1 (1-indexed, not 0-indexed)
    start_offset = (page_number * page_size) + 1 if page_number > 0 else 1

    request_body = f"""<ServiceRequest>
    <preferences>
        <limitResults>{page_size}</limitResults>
        <startFromOffset>{start_offset}</startFromOffset>
    </preferences>
</ServiceRequest>"""

    try:
        tag_response = requests.post(
            tag_url,
            headers=tag_headers,
            data=request_body,
            auth=HTTPBasicAuth(username, password),
            timeout=60
        )
    except requests.exceptions.Timeout:
        print("ERROR: Tag fetch request timed out after 60 seconds.")
        print("Please check your network connection and try again.")
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Network error during tag fetch: {e}")
        sys.exit(1)

    # Check for rate limiting (HTTP 429)
    if tag_response.status_code == 429:
        print("\n" + "="*70)
        print("RATE LIMIT REACHED (Tag List Endpoint)")
        print("="*70)
        print("Qualys API rate limit has been reached (300 calls/hour).")
        print("\nThe script cannot continue at this time.")
        print("\nTo retry:")
        print("1. Wait for the rate limit window to reset (typically 1 hour)")
        print("2. Run this script again")
        print("="*70)
        sys.exit(1)

    if tag_response.status_code != 200:
        print(f"\nFailed to fetch tags (HTTP {tag_response.status_code}).")
        print("Response from server:")
        print(f"Body: {tag_response.text or '(empty)'}")
        print(f"\nFull response content: {tag_response.content}")

        # Try to parse as XML to get error details
        if tag_response.text:
            try:
                error_root = ET.fromstring(tag_response.text)
                print(f"Parsed XML error: {ET.tostring(error_root, encoding='unicode')}")
            except:
                pass
        sys.exit(1)

    # Parse XML response
    try:
        root = ET.fromstring(tag_response.text)
    except ET.ParseError as e:
        print(f"\nERROR: Received invalid XML from API: {e}")
        print(f"Response text (first 500 chars): {tag_response.text[:500]}")
        sys.exit(1)

    # Extract tags from XML response
    # Expected structure: <ServiceResponse><data><Tag>...</Tag></data></ServiceResponse>
    tags = []
    data_element = root.find('.//data')

    if data_element is not None:
        # Find all Tag elements
        tag_elements = data_element.findall('Tag')

        for tag_elem in tag_elements:
            tag_dict = {}
            # Extract all child elements of the Tag
            for child in tag_elem:
                value = child.text
                # Decode HTML entities in tag names
                if child.tag == 'name' and value:
                    value = html.unescape(value)
                tag_dict[child.tag] = value
            tags.append(tag_dict)

    # If no tags returned, we're done
    if not tags:
        break

    all_tags.extend(tags)

    # If we got fewer tags than requested, we've reached the end
    if len(tags) < page_size:
        break

    page_number += 1

print(f"\nTotal tags found: {len(all_tags)}")

# ============================================================================
# Fetch detailed information for ALL tags using GET endpoint
# ============================================================================

if all_tags:
    print("\n" + "="*80)
    print(f"Generating detailed report ({len(all_tags)} tags)")
    print("="*80)

    # Collect data for reports
    if resume_from_progress:
        report_data = existing_progress.get('processed_tags', [])
        processed_tag_ids = {item['Tag ID'] for item in report_data}
        print(f"\nResuming from {len(report_data)} previously processed tags...")
    else:
        report_data = []
        processed_tag_ids = set()

    print("\nPress Ctrl+C to pause and resume later\n")

    # Process ALL tags with progress bar
    for idx, tag in enumerate(all_tags, 1):
        if interrupted:
            break
        tag_id = tag.get("id")
        if not tag_id:
            continue

        # Skip if already processed
        if tag_id in processed_tag_ids:
            continue

        # Progress bar
        progress_percent = (idx / len(all_tags)) * 100
        bar_length = 50
        filled_length = int(bar_length * idx // len(all_tags))
        bar = '█' * filled_length + '-' * (bar_length - filled_length)
        print(f'\r[{bar}] {progress_percent:.1f}% ({idx}/{len(all_tags)}) tags processed', end='', flush=True)

        # Construct the GET URL for this specific tag
        tag_detail_url = f"{qualys_api_url}/qps/rest/2.0/get/am/tag/{tag_id}"

        try:
            detail_response = requests.get(
                tag_detail_url,
                auth=HTTPBasicAuth(username, password),
                timeout=60
            )

            # Check for rate limiting (HTTP 429)
            if detail_response.status_code == 429:
                print("\n\n" + "="*70)
                print("RATE LIMIT REACHED (Tag Detail Endpoint)")
                print("="*70)
                print("Qualys API rate limit has been reached (300 calls/hour).")
                save_progress(report_data, platform, username)
                print("\nTo resume:")
                print("1. Wait for the rate limit window to reset (typically 1 hour)")
                print("2. Run this script again")
                print("3. Choose 'yes' when asked to resume from saved progress")
                print("="*70)
                sys.exit(0)

            if detail_response.status_code != 200:
                print(f"  ERROR: Failed to fetch details (HTTP {detail_response.status_code})")
                continue

            # Parse the XML response
            try:
                detail_root = ET.fromstring(detail_response.text)

                # Navigate to the Tag element
                tag_element = detail_root.find('.//Tag')
                if tag_element is None:
                    print(f"  ERROR: No Tag element found in response")
                    continue

                # Extract tag information
                tag_name = tag_element.findtext('name', 'N/A')
                if tag_name != 'N/A':
                    # Decode HTML entities like &lt; and &gt;
                    tag_name = html.unescape(tag_name)

                # Get parent tag name
                parent_tag_id = tag_element.findtext('parentTagId', '')
                if parent_tag_id:
                    # Fetch parent tag name by ID
                    parent_name = "N/A"
                    for t in all_tags:
                        if t.get("id") == parent_tag_id:
                            parent_name = t.get("name", "N/A")
                            break
                else:
                    parent_name = "-"

                # Count child tags
                children_element = tag_element.find('.//children/list')
                if children_element is not None:
                    child_count = len(children_element.findall('TagSimple'))
                else:
                    child_count = 0

                # Determine tag type (static vs dynamic)
                rule_type = tag_element.findtext('ruleType', '')
                if rule_type:
                    tag_type = 'Dynamic'
                    rule_type_display = rule_type

                    # Extract rule text and decode HTML entities
                    rule_text = tag_element.findtext('ruleText', 'N/A')
                    if rule_text != 'N/A':
                        # Decode HTML entities like &lt; and &gt;
                        rule_text = html.unescape(rule_text)

                        # For ASSET_SEARCH and NETWORK_RANGE_ENHANCED, remove XML declaration
                        if rule_type in ['ASSET_SEARCH', 'NETWORK_RANGE_ENHANCED']:
                            rule_text = rule_text.replace('<?xml version="1.0" encoding="UTF-8"?>', '').strip()

                        # Truncate rule text for table display
                        rule_text = rule_text.replace('\n', ' ').replace('\r', '')[:47] + '...'
                    else:
                        rule_text = 'N/A'
                else:
                    tag_type = 'Static'
                    rule_type_display = 'N/A'
                    rule_text = 'N/A'

                # Get ACS (Asset Criticality Score)
                acs = tag_element.findtext('criticalityScore', '-')
                if not acs or acs == 'N/A':
                    acs = '-'

                # Get created and modified dates
                created_date_raw = tag_element.findtext('created', '')
                modified_date_raw = tag_element.findtext('modified', '')

                # Format dates from ISO format (2014-02-06T19:14:50Z) to DD-MM-YYYY HH:MM:SS
                def format_date(date_str):
                    if not date_str:
                        return 'N/A'
                    try:
                        # Parse ISO format and convert to desired format
                        dt = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%SZ')
                        return dt.strftime('%d-%m-%Y %H:%M:%S')
                    except (ValueError, AttributeError):
                        return 'N/A'

                created_date = format_date(created_date_raw)
                modified_date = format_date(modified_date_raw)

                # Get asset count for this tag using the count endpoint
                # Use JWT authentication with the gateway URL
                try:
                    count_url = f"{gateway_url}/rest/2.0/count/am/asset"
                    count_headers = {
                        "Accept": "application/json",
                        "Authorization": f"Bearer {jwt_token}",
                        "Content-Type": "application/json"
                    }
                    count_body = {
                        "filters": [
                            {
                                "field": "tags.name",
                                "operator": "EQUALS",
                                "value": tag_name
                            }
                        ]
                    }

                    count_response = requests.post(
                        count_url,
                        headers=count_headers,
                        json=count_body,
                        timeout=30
                    )

                    if count_response.status_code == 200:
                        count_data = count_response.json()
                        asset_count = count_data.get('count', 'N/A')
                    elif count_response.status_code == 429:
                        # Rate limit hit - save progress and exit
                        print("\n\n" + "="*70)
                        print("RATE LIMIT REACHED (Asset Count Endpoint)")
                        print("="*70)
                        print("Qualys API rate limit has been reached (300 calls/hour).")
                        save_progress(report_data, platform, username)
                        print("\nTo resume:")
                        print("1. Wait for the rate limit window to reset (typically 1 hour)")
                        print("2. Run this script again")
                        print("3. Choose 'yes' when asked to resume from saved progress")
                        print("="*70)
                        sys.exit(0)
                    else:
                        asset_count = 'N/A'
                except (requests.exceptions.RequestException, json.JSONDecodeError, KeyError):
                    asset_count = 'N/A'

                # Store full rule text for reports (before truncation)
                if rule_type:
                    # For dynamic tags, get the full rule text before truncation
                    rule_text_full = tag_element.findtext('ruleText', 'N/A')
                    if rule_text_full != 'N/A':
                        rule_text_full = html.unescape(rule_text_full)
                        # For ASSET_SEARCH and NETWORK_RANGE_ENHANCED, remove XML declaration
                        if rule_type in ['ASSET_SEARCH', 'NETWORK_RANGE_ENHANCED']:
                            rule_text_full = rule_text_full.replace('<?xml version="1.0" encoding="UTF-8"?>', '').strip()
                        # Keep newlines for proper formatting in reports
                    else:
                        rule_text_full = 'N/A'
                else:
                    rule_text_full = 'N/A'

                # Add to report data
                report_data.append({
                    'Tag ID': tag_id,
                    'Tag Name': tag_name,
                    'Parent Name': parent_name,
                    'Child Tags': child_count,
                    'Asset Count': asset_count,
                    'Tag Type': tag_type,
                    'ACS': acs,
                    'Rule Type': rule_type_display,
                    'Rule Text': rule_text_full,
                    'Created': created_date,
                    'Modified': modified_date
                })

                # Mark as processed
                processed_tag_ids.add(tag_id)

                # Auto-save progress every 10 tags
                if len(report_data) % 10 == 0:
                    save_progress(report_data, platform, username, silent=True)

            except ET.ParseError as e:
                print(f"  ERROR: Invalid XML response: {e}")

        except requests.exceptions.Timeout:
            print(f"\n  ERROR: Request timed out for tag {tag_id}")
            print("  Network may be slow. Progress has been saved.")
            save_progress(report_data, platform, username)
            sys.exit(1)
        except requests.exceptions.RequestException as e:
            print(f"\n  ERROR: Network error for tag {tag_id}: {e}")
            print("  Saving progress before exit...")
            save_progress(report_data, platform, username)
            sys.exit(1)

    # Handle interruption
    if interrupted:
        save_progress(report_data, platform, username)
        print("\n\nScript paused. Run again and choose 'yes' to resume.")
        sys.exit(0)

    print("\n\n" + "="*80)
    print(f"Processing complete! {len(report_data)} tags processed.")
    print("="*80)

    # Generate Excel and HTML reports
    if report_data:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_username = "".join(c for c in username if c.isalnum() or c in ('-', '_')).lower()

        # ============================================================================
        # Excel Report Generation
        # ============================================================================
        excel_filename = f"tag_report_{platform}_{safe_username}_{timestamp}.xlsx"
        excel_filepath = os.path.abspath(excel_filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Tag Report"

        # Define colors
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = ['Tag Name', 'Parent Name', 'Child Tags', 'Asset Count', 'Tag Type', 'ACS', 'Rule Type', 'Rule Text', 'Created', 'Modified']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Helper function to sanitize cell values
        def sanitize_for_excel(value):
            if value is None:
                return ""
            value = str(value)
            # Remove illegal characters but preserve newline (0x0A) and carriage return (0x0D)
            illegal_chars = list(range(0x00, 0x09)) + [0x0B, 0x0C] + list(range(0x0E, 0x1F)) + list(range(0x7F, 0xA0))
            for char_code in illegal_chars:
                value = value.replace(chr(char_code), '')
            return value

        # Write data rows
        for row_num, row_data in enumerate(report_data, 2):
            ws.cell(row=row_num, column=1, value=sanitize_for_excel(row_data['Tag Name']))
            ws.cell(row=row_num, column=2, value=sanitize_for_excel(row_data['Parent Name']))
            ws.cell(row=row_num, column=3, value=row_data['Child Tags'])
            ws.cell(row=row_num, column=4, value=row_data['Asset Count'])
            ws.cell(row=row_num, column=5, value=sanitize_for_excel(row_data['Tag Type']))
            ws.cell(row=row_num, column=6, value=sanitize_for_excel(row_data['ACS']))
            ws.cell(row=row_num, column=7, value=sanitize_for_excel(row_data['Rule Type']))
            # Enable text wrapping for Rule Text column to preserve newlines
            rule_text_cell = ws.cell(row=row_num, column=8, value=sanitize_for_excel(row_data['Rule Text']))
            rule_text_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_num, column=9, value=sanitize_for_excel(row_data['Created']))
            ws.cell(row=row_num, column=10, value=sanitize_for_excel(row_data['Modified']))

        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            max_length = len(header)
            for row_data in report_data:
                cell_value = str(row_data.get(header, ''))
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save Excel file
        try:
            wb.save(excel_filename)
            excel_directory = os.path.dirname(excel_filepath) or os.getcwd()
            print(f"\nExcel report exported to {excel_filename}")
            print(f"File located at {excel_directory}\n")
        except (PermissionError, OSError) as e:
            print(f"\nWARNING: Failed to write Excel file: {e}\n")

        # ============================================================================
        # HTML Report Generation
        # ============================================================================
        html_filename = f"tag_report_{platform}_{safe_username}_{timestamp}.html"

        # Calculate statistics
        total_zero_asset_tags = sum(1 for row in report_data if row['Asset Count'] == 0 or row['Asset Count'] == '0')

        # Collect all unique rule types for filtering
        rule_types = set()
        for row in report_data:
            rule_type = row['Rule Type']
            if rule_type and rule_type != 'N/A':
                rule_types.add(rule_type)
        rule_types = sorted(list(rule_types))

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tag Report - {timestamp}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #224F91 0%, #1a3d73 100%);
            padding: 20px;
            min-height: 100vh;
        }}

        .container {{
            max-width: 1600px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: visible;
        }}

        .header {{
            background: linear-gradient(135deg, #34495E 0%, #2C3E50 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}

        .header .metadata {{
            font-size: 14px;
            opacity: 0.9;
            margin-top: 10px;
        }}

        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
        }}

        .summary-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
        }}

        .summary-card .number {{
            font-size: 36px;
            font-weight: bold;
            color: #224F91;
            margin-bottom: 5px;
        }}

        .summary-card .label {{
            font-size: 14px;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .controls {{
            padding: 20px 30px;
            background: white;
            border-bottom: 1px solid #e9ecef;
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
        }}

        .search-box {{
            flex: 1;
            min-width: 250px;
            position: relative;
        }}

        .search-box input {{
            width: 100%;
            padding: 10px 15px 10px 40px;
            border: 2px solid #e9ecef;
            border-radius: 6px;
            font-size: 14px;
            transition: border-color 0.3s;
        }}

        .search-box input:focus {{
            outline: none;
            border-color: #224F91;
        }}

        .search-box::before {{
            content: "🔍";
            position: absolute;
            left: 12px;
            top: 50%;
            transform: translateY(-50%);
        }}

        .table-container {{
            padding: 30px;
            overflow: visible;
        }}

        table {{
            width: 100%;
            min-width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            background: white;
            table-layout: fixed;
        }}

        thead {{
            background: #34495E;
            color: white;
        }}

        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            cursor: pointer;
            user-select: none;
            white-space: nowrap;
            position: -webkit-sticky;
            position: sticky;
            top: 0;
            z-index: 100;
            background: #34495E;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
            border-bottom: 2px solid #2C3E50;
        }}

        /* Ensure sticky works in all browsers */
        thead tr {{
            position: relative;
        }}

        th.sortable::after {{
            content: " ⇅";
            opacity: 0.5;
            font-size: 10px;
        }}

        th .resizer {{
            position: absolute;
            top: 0;
            right: 0;
            width: 5px;
            cursor: col-resize;
            user-select: none;
            height: 100%;
            z-index: 1;
        }}

        th .resizer:hover {{
            background-color: #224F91;
        }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #f0f0f0;
            font-size: 14px;
            word-wrap: break-word;
            overflow-wrap: break-word;
            vertical-align: top;
        }}

        /* Preserve newlines and formatting in rule text column */
        td:nth-child(8) {{
            white-space: pre-wrap;
        }}

        .tag-name-cell {{
            cursor: pointer;
            position: relative;
        }}

        .hierarchy-icon {{
            display: inline-block;
            width: 16px;
            margin-right: 5px;
            font-size: 12px;
            transition: transform 0.2s;
            cursor: pointer;
            user-select: none;
        }}

        .hierarchy-icon.expanded {{
            transform: rotate(90deg);
        }}

        .child-row {{
            display: none;
        }}

        .child-row.visible {{
            display: table-row;
        }}

        .child-indent {{
            padding-left: 40px;
        }}

        .child-indent-2 {{
            padding-left: 60px;
        }}

        .child-indent-3 {{
            padding-left: 80px;
        }}

        /* Set initial column widths */
        th:nth-child(1), td:nth-child(1) {{ width: 15%; }}  /* Tag Name */
        th:nth-child(2), td:nth-child(2) {{ width: 10%; }}  /* Parent Name */
        th:nth-child(3), td:nth-child(3) {{ width: 7%; }}   /* Child Tags */
        th:nth-child(4), td:nth-child(4) {{ width: 8%; }}   /* Asset Count */
        th:nth-child(5), td:nth-child(5) {{ width: 7%; }}   /* Tag Type */
        th:nth-child(6), td:nth-child(6) {{ width: 5%; }}   /* ACS */
        th:nth-child(7), td:nth-child(7) {{ width: 10%; }}  /* Rule Type */
        th:nth-child(8), td:nth-child(8) {{ width: 20%; }}  /* Rule Text */
        th:nth-child(9), td:nth-child(9) {{ width: 9%; }}   /* Created */
        th:nth-child(10), td:nth-child(10) {{ width: 9%; }} /* Modified */

        tbody tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}

        tbody tr:hover {{
            background-color: #e9ecef;
            transition: background-color 0.2s;
        }}

        .no-results {{
            text-align: center;
            padding: 40px;
            color: #6c757d;
            font-size: 16px;
        }}

        .btn {{
            padding: 10px 20px;
            background: #224F91;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            transition: background 0.3s;
        }}

        .btn:hover {{
            background: #1a3d73;
        }}

        .filter-container {{
            position: relative;
            display: inline-block;
        }}

        .filter-menu {{
            display: none;
            position: absolute;
            top: 100%;
            right: 0;
            background: white;
            border: 2px solid #e9ecef;
            border-radius: 6px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            z-index: 1000;
            min-width: 220px;
            margin-top: 5px;
            max-height: 400px;
            overflow-y: auto;
        }}

        .filter-menu.show {{
            display: block;
        }}

        .filter-menu label {{
            display: block;
            padding: 10px 15px;
            cursor: pointer;
            transition: background 0.2s;
            user-select: none;
        }}

        .filter-menu label:hover {{
            background: #f8f9fa;
        }}

        .filter-menu input[type="checkbox"] {{
            margin-right: 10px;
            cursor: pointer;
        }}

        .filter-section {{
            border-bottom: 1px solid #e9ecef;
            padding: 5px 0;
        }}

        .filter-section:last-child {{
            border-bottom: none;
        }}

        .filter-section-title {{
            padding: 8px 15px;
            font-weight: 600;
            font-size: 12px;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            background: #f8f9fa;
        }}

        .filter-subsection {{
            padding-left: 15px;
        }}

        .column-toggle-container {{
            position: relative;
            display: inline-block;
        }}

        .column-toggle-menu {{
            display: none;
            position: absolute;
            top: 100%;
            right: 0;
            background: white;
            border: 2px solid #e9ecef;
            border-radius: 6px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            z-index: 1000;
            min-width: 200px;
            margin-top: 5px;
        }}

        .column-toggle-menu.show {{
            display: block;
        }}

        .column-toggle-menu label {{
            display: block;
            padding: 10px 15px;
            cursor: pointer;
            transition: background 0.2s;
            user-select: none;
        }}

        .column-toggle-menu label:hover {{
            background: #f8f9fa;
        }}

        .column-toggle-menu input[type="checkbox"] {{
            margin-right: 10px;
            cursor: pointer;
        }}

        @media print {{
            body {{
                background: white;
                padding: 0;
            }}

            .controls, .btn {{
                display: none;
            }}

            .container {{
                box-shadow: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Qualys Tag Report</h1>
            <div class="metadata">
                User: {username} | Platform: {platform} | Generated: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}
            </div>
        </div>

        <div class="summary">
            <div class="summary-card">
                <div class="number">{len(report_data)}</div>
                <div class="label">Total Tags</div>
            </div>
            <div class="summary-card">
                <div class="number">{total_zero_asset_tags}</div>
                <div class="label">Zero Asset Tags</div>
            </div>
        </div>

        <div class="controls">
            <div class="search-box">
                <input type="text" id="searchInput" placeholder="Search tags...">
            </div>
            <button class="btn" id="expandAllBtn">Expand All</button>
            <button class="btn" id="collapseAllBtn">Collapse All</button>
            <div class="filter-container">
                <button class="btn" id="filterBtn">Filter Tags ▼</button>
                <div class="filter-menu" id="filterMenu">
                    <div class="filter-section">
                        <label><input type="checkbox" class="tag-type-filter" data-type="Static" checked> Static Tags</label>
                    </div>
                    <div class="filter-section">
                        <div class="filter-section-title">Dynamic Tags</div>
                        <label><input type="checkbox" class="tag-type-filter" data-type="Dynamic" id="dynamicAllCheckbox" checked> All Dynamic</label>
                        <div class="filter-subsection">"""

        # Add checkboxes for each dynamic rule type
        for rule_type in rule_types:
            html_content += f"""
                            <label><input type="checkbox" class="rule-type-filter" data-rule-type="{html.escape(rule_type)}" checked> {html.escape(rule_type)}</label>"""

        html_content += """
                        </div>
                    </div>
                </div>
            </div>
            <div class="column-toggle-container">
                <button class="btn" id="columnToggleBtn">Show/Hide Columns ▼</button>
                <div class="column-toggle-menu" id="columnToggleMenu">
                    <label><input type="checkbox" class="column-toggle" data-column="0" checked> Tag Name</label>
                    <label><input type="checkbox" class="column-toggle" data-column="1" checked> Parent Name</label>
                    <label><input type="checkbox" class="column-toggle" data-column="2" checked> Child Tags</label>
                    <label><input type="checkbox" class="column-toggle" data-column="3" checked> Asset Count</label>
                    <label><input type="checkbox" class="column-toggle" data-column="4" checked> Tag Type</label>
                    <label><input type="checkbox" class="column-toggle" data-column="5" checked> ACS</label>
                    <label><input type="checkbox" class="column-toggle" data-column="6" checked> Rule Type</label>
                    <label><input type="checkbox" class="column-toggle" data-column="7" checked> Rule Text</label>
                    <label><input type="checkbox" class="column-toggle" data-column="8" checked> Created</label>
                    <label><input type="checkbox" class="column-toggle" data-column="9" checked> Modified</label>
                </div>
            </div>
        </div>

        <div class="table-container">
            <table id="dataTable">
                <thead>
                    <tr>
                        <th data-column="0" class="sortable">Tag Name<div class="resizer"></div></th>
                        <th data-column="1" class="sortable">Parent Name<div class="resizer"></div></th>
                        <th data-column="2" class="sortable">Child Tags<div class="resizer"></div></th>
                        <th data-column="3" class="sortable">Asset Count<div class="resizer"></div></th>
                        <th data-column="4" class="sortable">Tag Type<div class="resizer"></div></th>
                        <th data-column="5" class="sortable">ACS<div class="resizer"></div></th>
                        <th data-column="6" class="sortable">Rule Type<div class="resizer"></div></th>
                        <th>Rule Text<div class="resizer"></div></th>
                        <th data-column="8" class="sortable">Created<div class="resizer"></div></th>
                        <th data-column="9" class="sortable">Modified<div class="resizer"></div></th>
                    </tr>
                </thead>
                <tbody>
"""

        # Build parent-child mapping for hierarchy
        tag_children = {}  # Map tag name -> list of child tag names
        for row_data in report_data:
            parent = row_data['Parent Name']
            if parent != '-':
                if parent not in tag_children:
                    tag_children[parent] = []
                tag_children[parent].append(row_data['Tag Name'])

        # Sort report_data so parents appear before their children
        def get_tag_by_name(name):
            for row in report_data:
                if row['Tag Name'] == name:
                    return row
            return None

        def sort_hierarchically(data):
            # Build a map for quick lookup
            tag_map = {row['Tag Name']: row for row in data}

            # Separate root tags (no parent or parent not found) and child tags
            root_tags = []
            child_tags = []

            for row in data:
                parent_name = row['Parent Name']
                if parent_name == '-' or parent_name not in tag_map:
                    root_tags.append(row)
                else:
                    child_tags.append(row)

            # Recursive function to add tag and its children
            def add_with_children(tag, result):
                result.append(tag)
                tag_name = tag['Tag Name']
                if tag_name in tag_children:
                    for child_name in tag_children[tag_name]:
                        child_tag = tag_map.get(child_name)
                        if child_tag:
                            add_with_children(child_tag, result)

            # Build the sorted list
            sorted_data = []
            for root in root_tags:
                add_with_children(root, sorted_data)

            return sorted_data

        report_data = sort_hierarchically(report_data)

        # Calculate hierarchy depth for each tag
        def get_depth(tag_name, depth=0):
            for row in report_data:
                if row['Tag Name'] == tag_name:
                    parent = row['Parent Name']
                    if parent == '-':
                        return depth
                    return get_depth(parent, depth + 1)
            return depth

        # Add table rows with HTML escaping and hierarchy data
        for row_data in report_data:
            tag_name = row_data['Tag Name']
            parent_name = row_data['Parent Name']
            child_count = row_data['Child Tags']
            is_parent = child_count > 0
            is_child = parent_name != '-'
            depth = get_depth(tag_name) if is_child else 0

            # Determine row classes
            row_classes = []
            if is_child:
                row_classes.append('child-row')

            row_class_str = f' class="{" ".join(row_classes)}"' if row_classes else ''

            # Build data attributes
            data_attrs = f'data-tag-name="{html.escape(tag_name)}"'
            if parent_name != '-':
                data_attrs += f' data-parent-name="{html.escape(parent_name)}"'
            data_attrs += f' data-depth="{depth}"'
            data_attrs += f' data-tag-type="{html.escape(row_data["Tag Type"])}"'
            data_attrs += f' data-rule-type="{html.escape(row_data["Rule Type"])}"'

            # Truncate rule text to 20 lines for HTML display
            rule_text_display = row_data['Rule Text']
            rule_text_full = str(rule_text_display)
            rule_text_full_escaped = html.escape(rule_text_full)

            # Truncate if rule text is not N/A
            if rule_text_display != 'N/A' and rule_text_display:
                text_str = str(rule_text_display)
                lines = text_str.split('\n')

                # Case 1: Multiple lines - truncate to 20 lines
                if len(lines) > 20:
                    truncated_lines = lines[:20]
                    rule_text_display = '\n'.join(truncated_lines) + '\n\n... [Truncated - hover to see full text]'
                # Case 2: Single long line - truncate to 500 characters
                elif len(lines) == 1 and len(text_str) > 500:
                    rule_text_display = text_str[:500] + '... [Truncated - hover to see full text]'

            rule_text_display_escaped = html.escape(str(rule_text_display))

            # Build tag name cell with hierarchy icon and indentation
            base_indent = 20 + (depth * 20)
            if is_parent:
                # Parent tags: cell has padding for depth, icon is absolute positioned, text has margin for icon
                cell_style = f'padding-left: {base_indent}px; position: relative;'
                tag_name_content = f'<span class="hierarchy-icon" data-tag="{html.escape(tag_name)}" style="position:absolute;left:{base_indent}px;">▶</span><span style="padding-left:21px;">{html.escape(tag_name)}</span>'
                tag_name_class = f' class="tag-name-cell" style="{cell_style}"'
            else:
                # Child tags without children: just indent with padding-left for both icon space and depth
                cell_style = f'padding-left: {base_indent + 21}px;'
                tag_name_content = f'{html.escape(tag_name)}'
                tag_name_class = f' style="{cell_style}"'

            html_content += f"""                    <tr{row_class_str} {data_attrs}>
                        <td{tag_name_class}>{tag_name_content}</td>
                        <td>{html.escape(str(row_data['Parent Name']))}</td>
                        <td>{html.escape(str(row_data['Child Tags']))}</td>
                        <td>{html.escape(str(row_data['Asset Count']))}</td>
                        <td>{html.escape(str(row_data['Tag Type']))}</td>
                        <td>{html.escape(str(row_data['ACS']))}</td>
                        <td>{html.escape(str(row_data['Rule Type']))}</td>
                        <td title="{rule_text_full_escaped}" class="rule-text-cell">{rule_text_display_escaped}</td>
                        <td>{html.escape(str(row_data['Created']))}</td>
                        <td>{html.escape(str(row_data['Modified']))}</td>
                    </tr>
"""

        html_content += """                </tbody>
            </table>
            <div class="no-results" id="noResults" style="display: none;">
                No matching records found
            </div>
        </div>
    </div>

    <script>
        const searchInput = document.getElementById('searchInput');
        const table = document.getElementById('dataTable');
        const tbody = table.querySelector('tbody');
        const noResults = document.getElementById('noResults');

        // Hierarchy management
        const hierarchyState = {}; // Track expanded state of parent tags

        function toggleHierarchy(tagName) {
            const icon = document.querySelector(`.hierarchy-icon[data-tag="${tagName}"]`);
            const childRows = document.querySelectorAll(`tr[data-parent-name="${tagName}"]`);

            const isExpanded = hierarchyState[tagName] || false;
            hierarchyState[tagName] = !isExpanded;

            if (hierarchyState[tagName]) {
                // Expand
                icon.classList.add('expanded');
                childRows.forEach(row => {
                    row.classList.add('visible');
                    row.style.display = '';
                    const childTagName = row.getAttribute('data-tag-name');
                    // If this child is also a parent and was previously expanded, show its children too
                    if (hierarchyState[childTagName]) {
                        const grandchildRows = document.querySelectorAll(`tr[data-parent-name="${childTagName}"]`);
                        grandchildRows.forEach(gr => {
                            gr.classList.add('visible');
                            gr.style.display = '';
                        });
                    }
                });
            } else {
                // Collapse
                icon.classList.remove('expanded');
                // Hide all descendants recursively
                function hideDescendants(parentName) {
                    const children = document.querySelectorAll(`tr[data-parent-name="${parentName}"]`);
                    children.forEach(row => {
                        row.classList.remove('visible');
                        row.style.display = 'none';
                        const childName = row.getAttribute('data-tag-name');
                        hideDescendants(childName);
                    });
                }
                hideDescendants(tagName);
            }
        }

        // Add click handlers to hierarchy icons
        document.querySelectorAll('.hierarchy-icon').forEach(icon => {
            icon.addEventListener('click', function(e) {
                e.stopPropagation();
                const tagName = this.getAttribute('data-tag');
                toggleHierarchy(tagName);
            });
        });

        // Add click handlers to tag name cells
        document.querySelectorAll('.tag-name-cell').forEach(cell => {
            cell.addEventListener('click', function(e) {
                if (!e.target.classList.contains('hierarchy-icon')) {
                    const icon = this.querySelector('.hierarchy-icon');
                    if (icon) {
                        const tagName = icon.getAttribute('data-tag');
                        toggleHierarchy(tagName);
                    }
                }
            });
        });

        // Expand All button
        document.getElementById('expandAllBtn').addEventListener('click', function() {
            document.querySelectorAll('.hierarchy-icon').forEach(icon => {
                const tagName = icon.getAttribute('data-tag');
                if (!hierarchyState[tagName]) {
                    toggleHierarchy(tagName);
                }
            });
        });

        // Collapse All button
        document.getElementById('collapseAllBtn').addEventListener('click', function() {
            document.querySelectorAll('.hierarchy-icon').forEach(icon => {
                const tagName = icon.getAttribute('data-tag');
                if (hierarchyState[tagName]) {
                    toggleHierarchy(tagName);
                }
            });
        });

        // Initialize: hide all child rows on page load
        function initializeHierarchy() {
            const rows = tbody.getElementsByTagName('tr');
            for (let row of rows) {
                if (row.classList.contains('child-row')) {
                    row.style.display = 'none';
                }
            }
        }
        initializeHierarchy();

        searchInput.addEventListener('keyup', function() {
            const filter = this.value.toLowerCase();
            const rows = Array.from(tbody.getElementsByTagName('tr'));
            let visibleCount = 0;

            if (filter === '') {
                // No search - restore hierarchy state
                rows.forEach(row => {
                    if (row.classList.contains('child-row')) {
                        // Only show if marked as visible by hierarchy toggle
                        if (row.classList.contains('visible')) {
                            row.style.display = '';
                            visibleCount++;
                        } else {
                            row.style.display = 'none';
                        }
                    } else {
                        // Parent or top-level row - always show
                        row.style.display = '';
                        visibleCount++;
                    }
                });
            } else {
                // Search mode - find matches and their parents
                const matchingRows = new Set();
                rows.forEach(row => {
                    const text = row.textContent.toLowerCase();
                    if (text.includes(filter)) {
                        matchingRows.add(row);
                        // Also include parent rows in the chain
                        let parentName = row.getAttribute('data-parent-name');
                        while (parentName) {
                            const parentRow = rows.find(r => r.getAttribute('data-tag-name') === parentName);
                            if (parentRow) {
                                matchingRows.add(parentRow);
                                parentName = parentRow.getAttribute('data-parent-name');
                            } else {
                                break;
                            }
                        }
                    }
                });

                // Show matching rows, hide non-matching
                rows.forEach(row => {
                    if (matchingRows.has(row)) {
                        row.style.display = '';
                        visibleCount++;
                    } else {
                        row.style.display = 'none';
                    }
                });
            }

            if (visibleCount === 0) {
                table.style.display = 'none';
                noResults.style.display = 'block';
            } else {
                table.style.display = 'table';
                noResults.style.display = 'none';
            }
        });

        let sortDirections = [true, true, true, true, true, true, true, true, true, true];

        function sortTable(columnIndex) {
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const direction = sortDirections[columnIndex];

            // For Tag Name column (index 0), use hierarchical sorting
            if (columnIndex === 0) {
                sortHierarchically(rows, direction);
            } else {
                // For other columns, use simple flat sorting
                rows.sort((a, b) => {
                    const aValue = a.cells[columnIndex].textContent.trim();
                    const bValue = b.cells[columnIndex].textContent.trim();

                    const aNum = parseFloat(aValue);
                    const bNum = parseFloat(bValue);

                    if (!isNaN(aNum) && !isNaN(bNum)) {
                        return direction ? aNum - bNum : bNum - aNum;
                    }

                    return direction ?
                        aValue.localeCompare(bValue) :
                        bValue.localeCompare(aValue);
                });

                rows.forEach(row => tbody.appendChild(row));
            }

            sortDirections[columnIndex] = !direction;
        }

        function sortHierarchically(rows, direction) {
            // Build a map of parent-child relationships
            const childrenMap = new Map();
            const rootRows = [];

            rows.forEach(row => {
                const parentName = row.getAttribute('data-parent-name');
                if (parentName) {
                    if (!childrenMap.has(parentName)) {
                        childrenMap.set(parentName, []);
                    }
                    childrenMap.get(parentName).push(row);
                } else {
                    rootRows.push(row);
                }
            });

            // Sort root rows by tag name
            rootRows.sort((a, b) => {
                const aName = a.getAttribute('data-tag-name');
                const bName = b.getAttribute('data-tag-name');
                return direction ?
                    aName.localeCompare(bName) :
                    bName.localeCompare(aName);
            });

            // Recursively add rows with their children
            function addRowWithChildren(row) {
                tbody.appendChild(row);
                const tagName = row.getAttribute('data-tag-name');
                const children = childrenMap.get(tagName);
                if (children) {
                    // Sort children by tag name
                    children.sort((a, b) => {
                        const aName = a.getAttribute('data-tag-name');
                        const bName = b.getAttribute('data-tag-name');
                        return direction ?
                            aName.localeCompare(bName) :
                            bName.localeCompare(aName);
                    });
                    children.forEach(child => addRowWithChildren(child));
                }
            }

            rootRows.forEach(row => addRowWithChildren(row));
        }

        // Column resizing functionality (must be set up BEFORE click handlers)
        const resizers = document.querySelectorAll('.resizer');
        let currentResizer = null;
        let currentTh = null;
        let startX = 0;
        let startWidth = 0;
        let isResizing = false;
        let hasMoved = false;

        resizers.forEach(resizer => {
            resizer.addEventListener('mousedown', function(e) {
                e.stopPropagation();
                e.preventDefault();
                isResizing = true;
                hasMoved = false;
                currentResizer = resizer;
                currentTh = resizer.parentElement;
                startX = e.pageX;
                startWidth = currentTh.offsetWidth;

                document.addEventListener('mousemove', handleMouseMove);
                document.addEventListener('mouseup', handleMouseUp);
            });
        });

        function handleMouseMove(e) {
            if (currentResizer) {
                hasMoved = true;
                const width = startWidth + (e.pageX - startX);
                if (width > 50) {
                    currentTh.style.width = width + 'px';
                }
            }
        }

        function handleMouseUp() {
            currentResizer = null;
            currentTh = null;
            isResizing = false;
            document.removeEventListener('mousemove', handleMouseMove);
            document.removeEventListener('mouseup', handleMouseUp);

            setTimeout(() => {
                hasMoved = false;
            }, 10);
        }

        // Add click handlers to sortable headers only
        const headers = document.querySelectorAll('th.sortable[data-column]');
        headers.forEach(th => {
            th.addEventListener('click', function(e) {
                if (hasMoved) {
                    hasMoved = false;
                    return;
                }
                if (!e.target.classList.contains('resizer')) {
                    const columnIndex = parseInt(this.getAttribute('data-column'));
                    sortTable(columnIndex);
                }
            });
        });

        // Filter functionality
        const filterBtn = document.getElementById('filterBtn');
        const filterMenu = document.getElementById('filterMenu');
        const tagTypeFilters = document.querySelectorAll('.tag-type-filter');
        const ruleTypeFilters = document.querySelectorAll('.rule-type-filter');
        const dynamicAllCheckbox = document.getElementById('dynamicAllCheckbox');

        // Toggle filter dropdown menu
        filterBtn.addEventListener('click', function(e) {
            e.stopPropagation();
            filterMenu.classList.toggle('show');
        });

        // Close filter dropdown when clicking outside
        document.addEventListener('click', function(e) {
            if (!filterMenu.contains(e.target) && e.target !== filterBtn) {
                filterMenu.classList.remove('show');
            }
        });

        // Function to apply filters
        function applyFilters() {
            const rows = tbody.querySelectorAll('tr');
            let visibleCount = 0;

            // Get selected filters
            const selectedTagTypes = new Set();
            tagTypeFilters.forEach(checkbox => {
                if (checkbox.checked) {
                    selectedTagTypes.add(checkbox.getAttribute('data-type'));
                }
            });

            const selectedRuleTypes = new Set();
            ruleTypeFilters.forEach(checkbox => {
                if (checkbox.checked) {
                    selectedRuleTypes.add(checkbox.getAttribute('data-rule-type'));
                }
            });

            rows.forEach(row => {
                const tagType = row.getAttribute('data-tag-type');
                const ruleType = row.getAttribute('data-rule-type');
                let shouldShow = false;

                // Check if row matches selected filters
                if (tagType === 'Static' && selectedTagTypes.has('Static')) {
                    shouldShow = true;
                } else if (tagType === 'Dynamic') {
                    // For dynamic tags, check if their rule type is selected
                    if (selectedRuleTypes.has(ruleType)) {
                        shouldShow = true;
                    }
                }

                // Apply visibility based on filter AND hierarchy state
                if (shouldShow) {
                    // Only show if not a child, or if it's a child that's marked visible by hierarchy
                    if (!row.classList.contains('child-row') || row.classList.contains('visible')) {
                        row.style.display = '';
                        visibleCount++;
                    } else {
                        row.style.display = 'none';
                    }
                } else {
                    row.style.display = 'none';
                }
            });

            // Update no results message
            if (visibleCount === 0) {
                table.style.display = 'none';
                noResults.style.display = 'block';
            } else {
                table.style.display = 'table';
                noResults.style.display = 'none';
            }
        }

        // Handle tag type filter changes
        tagTypeFilters.forEach(checkbox => {
            checkbox.addEventListener('change', function() {
                applyFilters();
            });
        });

        // Handle rule type filter changes
        ruleTypeFilters.forEach(checkbox => {
            checkbox.addEventListener('change', function() {
                // Update "All Dynamic" checkbox state
                const allRuleTypesChecked = Array.from(ruleTypeFilters).every(cb => cb.checked);
                const anyRuleTypeChecked = Array.from(ruleTypeFilters).some(cb => cb.checked);

                if (allRuleTypesChecked) {
                    dynamicAllCheckbox.checked = true;
                    dynamicAllCheckbox.indeterminate = false;
                } else if (anyRuleTypeChecked) {
                    dynamicAllCheckbox.checked = false;
                    dynamicAllCheckbox.indeterminate = true;
                } else {
                    dynamicAllCheckbox.checked = false;
                    dynamicAllCheckbox.indeterminate = false;
                }

                applyFilters();
            });
        });

        // Handle "All Dynamic" checkbox
        dynamicAllCheckbox.addEventListener('change', function() {
            const isChecked = this.checked;
            ruleTypeFilters.forEach(checkbox => {
                checkbox.checked = isChecked;
            });
            applyFilters();
        });

        // Column visibility toggle functionality
        const columnToggleBtn = document.getElementById('columnToggleBtn');
        const columnToggleMenu = document.getElementById('columnToggleMenu');
        const columnToggles = document.querySelectorAll('.column-toggle');

        // Toggle dropdown menu
        columnToggleBtn.addEventListener('click', function(e) {
            e.stopPropagation();
            columnToggleMenu.classList.toggle('show');
        });

        // Close dropdown when clicking outside
        document.addEventListener('click', function(e) {
            if (!columnToggleMenu.contains(e.target) && e.target !== columnToggleBtn) {
                columnToggleMenu.classList.remove('show');
            }
        });

        // Handle column visibility
        columnToggles.forEach(checkbox => {
            checkbox.addEventListener('change', function() {
                const columnIndex = parseInt(this.getAttribute('data-column'));
                const isVisible = this.checked;

                // Get all th and td elements for this column
                const ths = table.querySelectorAll('thead th');
                const trs = table.querySelectorAll('tbody tr');

                // Toggle header visibility
                if (ths[columnIndex]) {
                    ths[columnIndex].style.display = isVisible ? '' : 'none';
                }

                // Toggle cell visibility for all rows
                trs.forEach(tr => {
                    if (tr.cells[columnIndex]) {
                        tr.cells[columnIndex].style.display = isVisible ? '' : 'none';
                    }
                });
            });
        });
    </script>
</body>
</html>
"""

        # Write HTML file
        try:
            with open(html_filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            print(f"HTML report exported to {html_filename}")
            print(f"File located at {excel_directory}\n")
        except (PermissionError, OSError) as e:
            print(f"\nWARNING: Failed to write HTML file: {e}\n")

        # Successfully completed - delete progress file
        delete_progress(platform, username)

else:
    print("\nNo tags found in the subscription.")

# Logout / Invalidate token by posting with token=false
logout_data = {
    "username": username,
    "password": password,
    "token": "false"
}
try:
    logout_response = requests.post(auth_url, headers=auth_headers, data=logout_data, timeout=30)
    if logout_response.status_code == 200 or logout_response.status_code == 201:
        print("\nLogout successful.")
    else:
        print("\nLogout may have failed, but token will expire in 4 hours.")
except (requests.exceptions.Timeout, requests.exceptions.RequestException):
    print("\nLogout request failed, but token will expire in 4 hours.")
