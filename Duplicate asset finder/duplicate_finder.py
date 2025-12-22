"""
Qualys Duplicate Asset Finder

This script fetches Qualys host assets and identifies potential duplicates based on Asset Name, DNS Name, NetBIOS Name, MAC Address, and IPv4 Address. 

It generates both Excel and HTML reports with color-coded duplicate groups for easy review.

Command-line Arguments:
    --help                 : Show help message and exit
    --platform <PLATFORM>  : Qualys platform (US1, US2, US3, US4, UK, EU1, EU2, EU3, IN, CA, AE, AU, KSA)
    --username <USERNAME>  : Qualys username
    --password <PASSWORD>  : Qualys password (will prompt if not provided)
    --include-easm         : Include EASM assets in duplicate checking (overrides script default)
    --save-json            : Save filtered asset data to JSON file (overrides script default)

Example Usage:
    python duplicate_finder-v1.6.py --platform US1 --username user@example.com
    python duplicate_finder-v1.6.py --platform EU1 --username admin --password mypassword
    python duplicate_finder-v1.6.py --platform US1 --username user@example.com --include-easm --save-json
"""

import requests
import json
import getpass
import signal
import sys
import os
import argparse
from collections import defaultdict
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# ============================================================================
INCLUDE_EASM_ASSETS = False  # Change to True to include EASM assets by default
SAVE_JSON_OUTPUT = False     # Change to True to save asset data to JSON file
# ============================================================================

# Global flag for graceful shutdown
interrupted = False

def get_progress_filename(platform, username):
    """Generate progress filename based on platform and username"""
    # Sanitize username to be filesystem-safe
    safe_username = "".join(c for c in username if c.isalnum() or c in ('-', '_')).lower()
    return f"duplicate_finder_progress_{platform}_{safe_username}.json"

def save_progress(all_assets, last_seen_asset_id, platform, username, silent=False):
    """Save current progress to a JSON file (lightweight - only essential fields)"""
    progress_file = get_progress_filename(platform, username)

    # Extract only essential fields from each asset to reduce file size
    lightweight_assets = []
    for asset in all_assets:
        essential_asset = {
            "assetId": asset.get("assetId"),
            "assetName": asset.get("assetName"),
            "dnsName": asset.get("dnsName"),
            "netbiosName": asset.get("netbiosName"),
            "macAddress": asset.get("macAddress"),
            "address": asset.get("address"),
            "inventoryListData": asset.get("inventoryListData")
        }
        lightweight_assets.append(essential_asset)

    progress_data = {
        "platform": platform,
        "username": username,
        "last_seen_asset_id": last_seen_asset_id,
        "assets_fetched": len(lightweight_assets),
        "timestamp": datetime.now().isoformat(),  # ISO format for easier parsing
        "assets": lightweight_assets
    }

    try:
        with open(progress_file, 'w') as f:
            json.dump(progress_data, f)
        if not silent:
            print(f"\nProgress saved! ({len(lightweight_assets)} assets)")
    except (OSError, IOError) as e:
        print(f"\nWARNING: Failed to save progress: {e}")
        print("Continuing without progress save...")

def load_progress(platform, username):
    """Load progress from JSON file if it exists"""
    progress_file = get_progress_filename(platform, username)
    if os.path.exists(progress_file):
        try:
            with open(progress_file, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError, OSError, ValueError) as e:
            print(f"WARNING: Failed to load progress file: {e}")
            return None
    return None

def delete_progress(platform, username):
    """Delete the progress file"""
    progress_file = get_progress_filename(platform, username)
    if os.path.exists(progress_file):
        os.remove(progress_file)

def signal_handler(_sig, _frame):
    """Handle Ctrl+C gracefully"""
    global interrupted
    interrupted = True
    print("\n\nInterrupt received! Saving progress...")

# Register the signal handler for Ctrl+C
signal.signal(signal.SIGINT, signal_handler)

# Parse command-line arguments
parser = argparse.ArgumentParser(description='Qualys Duplicate Asset Finder')
parser.add_argument('--platform', type=str, help='Platform (US1, US2, US3, US4, UK, EU1, EU2, EU3, IN, CA, AE, AU, KSA)')
parser.add_argument('--username', type=str, help='Qualys username')
parser.add_argument('--password', type=str, help='Qualys password')
parser.add_argument('--include-easm', action='store_true', help='Include EASM assets in duplicate checking')
parser.add_argument('--save-json', action='store_true', help='Save filtered asset data to JSON file')
args = parser.parse_args()

# Override EASM inclusion flag if argument is provided
if args.include_easm:
    INCLUDE_EASM_ASSETS = True

# Override JSON save flag if argument is provided
if args.save_json:
    SAVE_JSON_OUTPUT = True

# Ask for the platform selection and username first (if not provided via args)
if args.platform:
    platform = args.platform.upper()
else:
    print("\nOptions: US1, US2, US3, US4, UK, EU1, EU2, EU3, IN, CA, AE, AU, KSA\n")
    platform = input("What platform is your account on? ").upper()

# Define gateway URLs for each platform (moved up for validation)
gateway_urls = {
    "US1": "https://gateway.qg1.apps.qualys.com",
    "US2": "https://gateway.qg2.apps.qualys.com",
    "US3": "https://gateway.qg3.apps.qualys.com",
    "US4": "https://gateway.qg4.apps.qualys.com",
    "UK": "https://gateway.qg1.apps.qualys.co.uk",
    "EU1": "https://gateway.qg1.apps.qualys.eu",
    "EU2": "https://gateway.qg2.apps.qualys.eu",
    "EU3": "https://gateway.qg3.apps.qualys.it",
    "IN": "https://gateway.qg1.apps.qualys.in",
    "CA": "https://gateway.qg1.apps.qualys.ca",
    "AE": "https://gateway.qg1.apps.qualys.ae",
    "AU": "https://gateway.qg1.apps.qualys.com.au",
    "KSA": "https://gateway.qg1.apps.qualysksa.com"
}

# Select the correct gateway URL
if platform in gateway_urls:
    gateway_url = gateway_urls[platform]
else:
    print("Invalid platform selection. Exiting")
    exit(1)

if args.username:
    username = args.username
else:
    username = input("Enter your username: ")

# Check for existing progress for THIS specific platform and username
existing_progress = load_progress(platform, username)
resume_from_progress = False

if existing_progress:
    # Check if progress is older than 24 hours
    try:
        progress_timestamp = datetime.fromisoformat(existing_progress.get('timestamp'))
        age_hours = (datetime.now() - progress_timestamp).total_seconds() / 3600

        if age_hours > 24:
            print("\n" + "="*70)
            print("STALE PROGRESS DETECTED")
            print("="*70)
            print(f"Progress file is {age_hours:.1f} hours old (saved: {existing_progress.get('timestamp')})")
            print("Progress older than 24 hours is considered stale and will be discarded.")
            print("Starting fresh session...")
            print("="*70)
            delete_progress(platform, username)
        else:
            print("\n" + "="*70)
            print("PREVIOUS SESSION DETECTED")
            print("="*70)
            print(f"Platform: {platform}")
            print(f"Username: {username}")
            print(f"Assets fetched: {existing_progress.get('assets_fetched')}")
            print(f"Last saved: {existing_progress.get('timestamp')} ({age_hours:.1f} hours ago)")
            print("="*70)

            resume_choice = input("\nDo you want to resume from saved progress? (yes/no): ").strip().lower()
            if resume_choice in ['yes', 'y']:
                resume_from_progress = True
                print(f"\nResuming session for {username} on {platform}...")
            else:
                print("\nStarting fresh session (deleting saved progress)...")
                delete_progress(platform, username)
    except (ValueError, TypeError):
        # If timestamp parsing fails, treat as stale
        print("\n" + "="*70)
        print("INVALID PROGRESS DETECTED")
        print("="*70)
        print("Progress file has invalid timestamp. Starting fresh session...")
        print("="*70)
        delete_progress(platform, username)

# Get password
if args.password:
    password = args.password
elif not resume_from_progress:
    password = getpass.getpass("Enter your password: ")
else:
    # When resuming, still need password for authentication
    password = getpass.getpass(f"Enter password for {username}: ")

# Define the authentication URL using the gateway URL
auth_url = f"{gateway_url}/auth"

# Authentication headers and data for JWT
auth_headers = {
    "Content-Type": "application/x-www-form-urlencoded"
}
auth_data = {
    "username": username,
    "password": password,
    "token": "true"
}

# Perform authentication to get JWT
try:
    auth_response = requests.post(auth_url, headers=auth_headers, data=auth_data, timeout=60)
except requests.exceptions.Timeout:
    print("ERROR: Authentication request timed out after 60 seconds.")
    print("Please check your network connection and try again.")
    exit(1)
except requests.exceptions.RequestException as e:
    print(f"ERROR: Network error during authentication: {e}")
    exit(1)

if auth_response.status_code != 201:
    print(f"Authentication failed. (HTTP {auth_response.status_code}).")
    print("Response from server:")
    print(auth_response.text or "No additional error details provided.")
    exit(1)

# Extract JWT token from response body (plain string)
jwt_token = auth_response.text.strip()
print("\nAuthentication successful.")

# Now fetch all assets using the asset endpoint
asset_url = f"{gateway_url}/rest/2.0/search/am/asset"
asset_headers = {
    "Accept": "application/json",
    "Authorization": f"Bearer {jwt_token}",
    "Content-Type": "application/json"
}

print("\nFetching assets...")
if resume_from_progress:
    print("(Press Ctrl+C at any time to pause and save progress)\n")
else:
    print("(Press Ctrl+C at any time to pause and save progress)\n")

# Initialize or restore from progress
if resume_from_progress:
    all_assets = existing_progress.get('assets', [])
    last_seen_asset_id = existing_progress.get('last_seen_asset_id')
    print(f"Resuming from {len(all_assets)} previously fetched assets...")
else:
    all_assets = []
    last_seen_asset_id = None

has_more = True
page_size = 300  # Max allowed by Qualys
api_call_count = 0  # Track number of API calls

while has_more and not interrupted:
    params = {"pageSize": page_size}
    if last_seen_asset_id:
        params["lastSeenAssetId"] = last_seen_asset_id

    try:
        asset_response = requests.post(asset_url, headers=asset_headers, params=params, json={}, timeout=60)
    except requests.exceptions.Timeout:
        print("\n" + "="*70)
        print("REQUEST TIMEOUT")
        print("="*70)
        print("Asset fetch request timed out after 60 seconds.")
        save_progress(all_assets, last_seen_asset_id, platform, username)
        print("\nProgress saved. Please check your network and try again.")
        print("="*70)
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        print(f"\nNetwork error during asset fetch: {e}")
        save_progress(all_assets, last_seen_asset_id, platform, username)
        sys.exit(1)

    api_call_count += 1

    # Check for rate limiting (HTTP 429)
    if asset_response.status_code == 429:
        print("\n" + "="*70)
        print("RATE LIMIT REACHED")
        print("="*70)
        print("Qualys API rate limit has been reached (300 calls/hour).")
        save_progress(all_assets, last_seen_asset_id, platform, username)
        print("\nTo resume:")
        print("1. Wait for the rate limit window to reset (typically 1 hour)")
        print("2. Run this script again")
        print("3. Choose 'yes' when asked to resume from saved progress")
        print("="*70)
        sys.exit(0)

    # HTTP 204 means "No Content" - all assets have been fetched
    if asset_response.status_code == 204:
        print("All assets fetched (no more data available).")
        has_more = False
        break

    if asset_response.status_code != 200:
        print(f"\nFailed to fetch assets (HTTP {asset_response.status_code}).")
        print("Response from server:")
        print(asset_response.text or "No additional error details provided.")
        # Save progress before exiting on error
        if len(all_assets) > 0:
            print("\nSaving progress before exit...")
            save_progress(all_assets, last_seen_asset_id, platform, username)
        sys.exit(1)

    # Parse JSON response with error handling
    try:
        data = asset_response.json()
    except json.JSONDecodeError as e:
        print(f"\nERROR: Received invalid JSON from API: {e}")
        print(f"Response text (first 500 chars): {asset_response.text[:500]}")
        if len(all_assets) > 0:
            print("\nSaving progress before exit...")
            save_progress(all_assets, last_seen_asset_id, platform, username)
        sys.exit(1)

    # Safely extract assets with type checking
    asset_list_data = data.get("assetListData")
    if isinstance(asset_list_data, dict):
        assets = asset_list_data.get("asset", [])
    else:
        assets = []
    all_assets.extend(assets)

    current_has_more = data.get("hasMore", 0) == 1
    current_last_seen = data.get("lastSeenAssetId")

    print(f"Fetched {len(all_assets)} assets so far...")

    # Auto-save progress after every fetch (lightweight checkpoint - silent)
    save_progress(all_assets, current_last_seen, platform, username, silent=True)

    # Show reminder every 10 API calls
    if api_call_count % 10 == 0:
        print("(Press Ctrl+C at any time to pause and save progress)")

    has_more = current_has_more
    last_seen_asset_id = current_last_seen

# Handle interruption
if interrupted:
    save_progress(all_assets, last_seen_asset_id, platform, username)
    print("\nScript paused. Run again and choose 'yes' to resume.")
    sys.exit(0)

# Check for potential duplicates across multiple fields
if INCLUDE_EASM_ASSETS:
    print("\nChecking for potential duplicates (including EASM assets)...\n")
    # Use all assets
    assets_to_check = all_assets
else:
    print("\nChecking for potential duplicates (ignoring assets with EASM as the only source)...\n")
    # Filter out assets where EASM is the only source
    assets_to_check = []
    for asset in all_assets:
        inventory_list = asset.get('inventoryListData')
        if inventory_list is None:
            # If there's no inventory list, include the asset (no EASM source to filter)
            assets_to_check.append(asset)
        else:
            # inventory_list is guaranteed to be not None here
            inventory_items = inventory_list.get('inventory', [])
            # Handle cases where source might be None
            sources = [item.get('source', '').upper() for item in inventory_items if item and item.get('source')]
            # Exclude only if EASM is the sole source (not if it has other sources too)
            if len(sources) == 0 or not (len(sources) == 1 and sources[0] == 'EASM'):
                assets_to_check.append(asset)

print(f"Total assets: {len(all_assets)}")
print(f"Assets to check for duplicates: {len(assets_to_check)}\n")

# Track flagged duplicate pairs (as frozensets of asset IDs) to avoid repeats
flagged_pairs = set()

# List to collect all duplicate assets for CSV export
csv_data = []
# Track which duplicate group each CSV row belongs to (for coloring)
csv_row_to_group = []
# Track which assets have already been added to avoid duplicates in the file
added_assets = set()

# Field display names and getters
fields = [
    ("assetName", "Asset name", lambda x: str(x.get("assetName") or "").strip().lower()),
    ("dnsName", "DNS name", lambda x: str(x.get("dnsName") or "").strip().lower()),
    ("netbiosName", "NetBIOS name", lambda x: str(x.get("netbiosName") or "").strip().lower()),
    ("macAddress", "MAC address", lambda x: str(x.get("macAddress") or "").strip().lower()),
    ("ipv4Address", "IPv4 address", lambda x: str(x.get("address") or "").strip())
]

for field_name, display_name, normalize_func in fields:
    groups = defaultdict(list)
    for asset in assets_to_check:
        value = normalize_func(asset)
        if value:  # Only consider non-empty values
            asset_id = asset.get("assetId")
            # Ensure assetId exists and is a valid identifier (not None, empty, or non-numeric)
            if asset_id is not None and asset_id != "":
                try:
                    # Ensure asset_id can be converted to int (Qualys asset IDs are integers)
                    asset_id = int(asset_id) if not isinstance(asset_id, int) else asset_id
                    groups[value].append({
                        "asset_id": asset_id,
                        "asset": asset
                    })
                except (ValueError, TypeError):
                    # Skip assets with invalid asset IDs
                    continue
    
    new_duplicates = []
    for value, group in groups.items():
        if len(group) > 1:
            # Check if this group has any new pairs not flagged
            group_ids = [item["asset_id"] for item in group]
            group_pairs = [frozenset([id1, id2]) for id1, id2 in combinations(group_ids, 2)]
            if all(pair not in flagged_pairs for pair in group_pairs):
                new_duplicates.append((value, group))
                # Flag all pairs in this group
                for pair in group_pairs:
                    flagged_pairs.add(pair)
    
    num_duplicates = len(new_duplicates)
    if field_name == "assetName":
        print(f"--------------------------------------------------------------------")
        print(f"\n{num_duplicates} Potential duplicates based on {display_name}")
    else:
        print(f"--------------------------------------------------------------------")
        print(f"\n{num_duplicates} Potential duplicates based on {display_name}, not reported previously\n")
    
    if num_duplicates > 0:
        for value, group in new_duplicates:
            print(f"\n- {display_name}: '{value}'")
            # Get current group number for coloring
            if len(csv_row_to_group) == 0:
                current_group_num = 0
            else:
                # Get the last group number and increment it
                current_group_num = csv_row_to_group[-1] + 1

            for item in group:
                asset = item['asset']
                asset_id = item['asset_id']

                # Extract all sources from inventoryListData
                inventory_list = asset.get('inventoryListData')
                if inventory_list and isinstance(inventory_list, dict):
                    inventory_items = inventory_list.get('inventory', [])
                else:
                    inventory_items = []

                if inventory_items and len(inventory_items) > 0:
                    # Handle cases where item or source might be None
                    sources = [item.get('source', 'Unknown') for item in inventory_items if item and isinstance(item, dict)]
                    source = ', '.join(sources) if sources else 'Unknown'
                    # Get the most recent lastUpdated timestamp from all inventory sources
                    last_updated_values = [item.get('lastUpdated') for item in inventory_items if item and isinstance(item, dict) and item.get('lastUpdated')]
                    last_activity_ms = max(last_updated_values) if last_updated_values else None
                else:
                    source = 'Unknown'
                    last_activity_ms = None

                # Extract fields for CSV
                address = asset.get('address', '')
                dns_name = asset.get('dnsName', '')
                asset_name = asset.get('assetName', '')

                # Convert last activity timestamp with validation
                if last_activity_ms:
                    try:
                        # Validate timestamp is reasonable (between 1970 and 2100)
                        # 4102444800000 ms = Jan 1, 2100
                        if isinstance(last_activity_ms, (int, float)) and 0 <= last_activity_ms <= 4102444800000:
                            last_activity_dt = datetime.fromtimestamp(last_activity_ms / 1000)
                            last_activity = last_activity_dt.strftime("%d-%m-%Y %H:%M:%S")
                        else:
                            last_activity = f'Invalid ({last_activity_ms})'
                    except (OSError, ValueError, OverflowError):
                        last_activity = 'Invalid timestamp'
                else:
                    last_activity = 'N/A'

                # Only add to CSV if this asset hasn't been added before
                if asset_id not in added_assets:
                    # Create the row data
                    row_data = {
                        'Asset ID': asset_id,
                        'Address': address,
                        'DNS Name': dns_name,
                        'Asset Name': asset_name,
                        'Source': source,
                        'Last Activity': last_activity
                    }
                    # Add both row data and group number atomically to maintain sync
                    csv_data.append(row_data)
                    csv_row_to_group.append(current_group_num)
                    # Mark this asset as added
                    added_assets.add(asset_id)

                    # Sanity check: ensure lists stay in sync
                    if len(csv_data) != len(csv_row_to_group):
                        print(f"WARNING: Data synchronization error detected. csv_data has {len(csv_data)} entries but csv_row_to_group has {len(csv_row_to_group)} entries.")
                        # Force sync by padding csv_row_to_group if needed
                        while len(csv_row_to_group) < len(csv_data):
                            csv_row_to_group.append(current_group_num)

                # Format output with asset name for non-assetName duplicates
                if field_name == "assetName":
                    # Don't include asset name in output as it would be redundant
                    print(f"  * Asset ID: {asset_id} | Last Activity: {last_activity} | (Source: {source})")
                else:
                    # Include asset name in output
                    asset_name_display = asset_name if asset_name else "(unavailable)"
                    print(f"  * Asset ID: {asset_id} | Asset name: {asset_name_display} | Last Activity: {last_activity} | (Source: {source})")

# Calculate total number of unique duplicate assets
unique_duplicate_assets = set()
for pair in flagged_pairs:
    unique_duplicate_assets.update(pair)
total_duplicates = len(unique_duplicate_assets)

print(f"\n--------------------------------------------------------------------")
print(f"\nTotal potential duplicate assets found: {total_duplicates}\n")

# Write duplicate assets to Excel file with alternating colors
if csv_data:
    import os
    from datetime import datetime

    # Final validation: ensure csv_data and csv_row_to_group are in sync
    if len(csv_data) != len(csv_row_to_group):
        print(f"\nERROR: Data synchronization mismatch detected before export!")
        print(f"csv_data has {len(csv_data)} entries but csv_row_to_group has {len(csv_row_to_group)} entries.")
        print("Attempting to fix by padding csv_row_to_group...")
        while len(csv_row_to_group) < len(csv_data):
            csv_row_to_group.append(0)  # Use group 0 as fallback

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Sanitize username for filename
    safe_username = "".join(c for c in username if c.isalnum() or c in ('-', '_')).lower()
    excel_filename = f"duplicate_assets_{platform}_{safe_username}_{timestamp}.xlsx"
    excel_filepath = os.path.abspath(excel_filename)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate Assets"

    # Define colors for alternating groups
    color1 = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")  # Medium light blue
    color2 = PatternFill(start_color="FAD7A0", end_color="FAD7A0", fill_type="solid")  # Light peach/orange

    # Header styling
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")  # Dark blue-gray
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    headers = ['Asset ID', 'Address', 'DNS Name', 'Asset Name', 'Source', 'Last Activity']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Helper function to sanitize cell values for Excel
    def sanitize_for_excel(value):
        """Remove illegal characters from cell values"""
        if value is None:
            return ""
        # Convert to string
        value = str(value)
        # Remove control characters (0x00-0x1F except tab, newline, carriage return)
        # Also remove 0x7F-0x9F
        illegal_chars = list(range(0x00, 0x09)) + [0x0B, 0x0C] + list(range(0x0E, 0x20)) + list(range(0x7F, 0xA0))
        for char_code in illegal_chars:
            value = value.replace(chr(char_code), '')
        return value

    # Pre-calculate column widths from csv_data (much faster than iterating through Excel cells)
    column_widths = [len(header) for header in headers]  # Start with header lengths
    for row_data in csv_data:
        column_widths[0] = max(column_widths[0], len(str(row_data['Asset ID'])))
        column_widths[1] = max(column_widths[1], len(str(row_data['Address'])))
        column_widths[2] = max(column_widths[2], len(str(row_data['DNS Name'])))
        column_widths[3] = max(column_widths[3], len(str(row_data['Asset Name'])))
        column_widths[4] = max(column_widths[4], len(str(row_data['Source'])))
        column_widths[5] = max(column_widths[5], len(str(row_data['Last Activity'])))

    # Apply column widths
    for col_num, width in enumerate(column_widths, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        adjusted_width = min(width + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Write data rows with alternating colors per group
    row_num = 2
    for idx, row_data in enumerate(csv_data):
        # Get the group number for this row
        group_num = csv_row_to_group[idx]

        # Alternate color based on group number
        row_fill = color1 if group_num % 2 == 0 else color2

        # Write row data with sanitized values
        ws.cell(row=row_num, column=1, value=sanitize_for_excel(row_data['Asset ID'])).fill = row_fill
        ws.cell(row=row_num, column=2, value=sanitize_for_excel(row_data['Address'])).fill = row_fill
        ws.cell(row=row_num, column=3, value=sanitize_for_excel(row_data['DNS Name'])).fill = row_fill
        ws.cell(row=row_num, column=4, value=sanitize_for_excel(row_data['Asset Name'])).fill = row_fill
        ws.cell(row=row_num, column=5, value=sanitize_for_excel(row_data['Source'])).fill = row_fill
        ws.cell(row=row_num, column=6, value=sanitize_for_excel(row_data['Last Activity'])).fill = row_fill

        row_num += 1

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save the workbook with error handling
    try:
        wb.save(excel_filename)
        excel_directory = os.path.dirname(excel_filepath) or os.getcwd()
        print(f"Duplicate assets exported to {excel_filename}\n")
        print(f"File located at {excel_directory}\n")
    except PermissionError:
        print(f"\nERROR: Permission denied when writing to {excel_filename}")
        print("The file may be open in Excel or you may not have write permissions.")
        print("Please close the file if it's open and try again, or check file permissions.\n")
        sys.exit(1)
    except OSError as e:
        print(f"\nERROR: Failed to write Excel file: {e}")
        print("Possible causes: disk full, invalid filename, or file system error.\n")
        sys.exit(1)
    except Exception as e:
        print(f"\nERROR: Unexpected error when saving Excel file: {e}\n")
        sys.exit(1)

    # Generate HTML report
    html_filename = f"duplicate_assets_{platform}_{safe_username}_{timestamp}.html"
    html_filepath = os.path.abspath(html_filename)

    # Calculate percentage safely before HTML generation
    duplicate_percentage = round((total_duplicates / len(all_assets) * 100), 1) if len(all_assets) > 0 else 0

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Duplicate Assets Report - {timestamp}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #224F91 0%, #1a3d73 100%);
            padding: 20px;
            min-height: 100vh;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: visible;
        }}

        .header {{
            background: linear-gradient(135deg, #34495E 0%, #2C3E50 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}

        .header .metadata {{
            font-size: 14px;
            opacity: 0.9;
            margin-top: 10px;
        }}

        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
        }}

        .summary-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
        }}

        .summary-card .number {{
            font-size: 36px;
            font-weight: bold;
            color: #224F91;
            margin-bottom: 5px;
        }}

        .summary-card .label {{
            font-size: 14px;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .controls {{
            padding: 20px 30px;
            background: white;
            border-bottom: 1px solid #e9ecef;
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
        }}

        .search-box {{
            flex: 1;
            min-width: 250px;
            position: relative;
        }}

        .search-box input {{
            width: 100%;
            padding: 10px 15px 10px 40px;
            border: 2px solid #e9ecef;
            border-radius: 6px;
            font-size: 14px;
            transition: border-color 0.3s;
        }}

        .search-box input:focus {{
            outline: none;
            border-color: #224F91;
        }}

        .search-box::before {{
            content: "🔍";
            position: absolute;
            left: 12px;
            top: 50%;
            transform: translateY(-50%);
        }}

        .table-container {{
            padding: 30px;
            overflow: visible;
        }}

        table {{
            width: 100%;
            min-width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            background: white;
            table-layout: fixed;
        }}

        thead {{
            background: #34495E;
            color: white;
        }}

        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            cursor: pointer;
            user-select: none;
            white-space: nowrap;
            position: -webkit-sticky;
            position: sticky;
            top: 0;
            z-index: 100;
            background: #34495E;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
            border-bottom: 2px solid #2C3E50;
        }}

        /* Ensure sticky works in all browsers */
        thead tr {{
            position: relative;
        }}

        th .resizer {{
            position: absolute;
            top: 0;
            right: 0;
            width: 5px;
            cursor: col-resize;
            user-select: none;
            height: 100%;
        }}

        th .resizer:hover {{
            background-color: #224F91;
        }}

        th:hover {{
            background: #2C3E50;
        }}

        th::after {{
            content: " ⇅";
            opacity: 0.5;
            font-size: 10px;
        }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #f0f0f0;
            font-size: 14px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}

        tr.group-0 {{
            background-color: #AED6F1;
        }}

        tr.group-1 {{
            background-color: #FAD7A0;
        }}

        tr:hover {{
            opacity: 0.8;
            transition: opacity 0.2s;
        }}

        .no-results {{
            text-align: center;
            padding: 40px;
            color: #6c757d;
            font-size: 16px;
        }}

        .footer {{
            text-align: center;
            padding: 20px;
            background: #f8f9fa;
            color: #6c757d;
            font-size: 13px;
            border-top: 1px solid #e9ecef;
        }}

        .legend {{
            display: flex;
            gap: 20px;
            align-items: center;
            font-size: 13px;
        }}

        .legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .legend-color {{
            width: 24px;
            height: 24px;
            border-radius: 4px;
            border: 1px solid #ddd;
        }}

        .btn {{
            padding: 10px 20px;
            background: #224F91;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            transition: background 0.3s;
        }}

        .btn:hover {{
            background: #1a3d73;
        }}

        @media print {{
            body {{
                background: white;
                padding: 0;
            }}

            .controls, .btn {{
                display: none;
            }}

            .container {{
                box-shadow: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Duplicate Assets Report</h1>
            <div class="metadata">
                User: {username} | Generated: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}
            </div>
        </div>

        <div class="summary">
            <div class="summary-card">
                <div class="number">{len(all_assets)}</div>
                <div class="label">Host Assets</div>
            </div>
            <div class="summary-card">
                <div class="number">{total_duplicates}</div>
                <div class="label">Potential Duplicates</div>
            </div>
            <div class="summary-card">
                <div class="number">{duplicate_percentage}%</div>
                <div class="label">Duplicate Ratio</div>
            </div>
        </div>

        <div class="controls">
            <div class="search-box">
                <input type="text" id="searchInput" placeholder="Search by Asset ID, Address, DNS Name, Asset Name, Source, or Last Activity...">
            </div>
            <button class="btn" id="resetSortBtn">Reset Sort</button>
            <button class="btn" onclick="window.print()">Print Report</button>
        </div>

        <div class="table-container">
            <table id="dataTable">
                <thead>
                    <tr>
                        <th data-column="0">Asset ID<div class="resizer"></div></th>
                        <th data-column="1">Address<div class="resizer"></div></th>
                        <th data-column="2">DNS Name<div class="resizer"></div></th>
                        <th data-column="3">Asset Name<div class="resizer"></div></th>
                        <th data-column="4">Source<div class="resizer"></div></th>
                        <th data-column="5">Last Activity<div class="resizer"></div></th>
                    </tr>
                </thead>
                <tbody>
"""

    # Add table rows with HTML escaping for security
    import html
    for idx, row_data in enumerate(csv_data):
        group_num = csv_row_to_group[idx]
        group_class = f"group-{group_num % 2}"

        html_content += f"""                    <tr class="{group_class}" data-original-index="{idx}">
                        <td>{html.escape(str(row_data['Asset ID']))}</td>
                        <td>{html.escape(str(row_data['Address']))}</td>
                        <td>{html.escape(str(row_data['DNS Name']))}</td>
                        <td>{html.escape(str(row_data['Asset Name']))}</td>
                        <td>{html.escape(str(row_data['Source']))}</td>
                        <td>{html.escape(str(row_data['Last Activity']))}</td>
                    </tr>
"""

    html_content += """                </tbody>
            </table>
            <div class="no-results" id="noResults" style="display: none;">
                No matching records found
            </div>
        </div>

        <div class="footer">
            Report contains potential duplicate assets based on Asset Name, DNS Name, NetBIOS Name, MAC Address, and IPv4 Address """

    # Add EASM status message
    if INCLUDE_EASM_ASSETS:
        html_content += "(EASM assets included)"
    else:
        html_content += "(assets with EASM as the only source excluded)"

    html_content += """
        </div>
    </div>

    <script>
        // Search functionality
        const searchInput = document.getElementById('searchInput');
        const table = document.getElementById('dataTable');
        const tbody = table.querySelector('tbody');
        const noResults = document.getElementById('noResults');
        const resetSortBtn = document.getElementById('resetSortBtn');

        // Store original row order on page load
        const originalRowOrder = Array.from(tbody.querySelectorAll('tr'));

        searchInput.addEventListener('keyup', function() {
            const filter = this.value.toLowerCase();
            const rows = tbody.getElementsByTagName('tr');
            let visibleCount = 0;

            for (let row of rows) {
                const text = row.textContent.toLowerCase();
                if (text.includes(filter)) {
                    row.style.display = '';
                    visibleCount++;
                } else {
                    row.style.display = 'none';
                }
            }

            if (visibleCount === 0) {
                table.style.display = 'none';
                noResults.style.display = 'block';
            } else {
                table.style.display = 'table';
                noResults.style.display = 'none';
            }
        });

        // Sort functionality
        let sortDirections = [true, true, true, true, true, true];

        function sortTable(columnIndex) {
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const direction = sortDirections[columnIndex];

            rows.sort((a, b) => {
                const aValue = a.cells[columnIndex].textContent.trim();
                const bValue = b.cells[columnIndex].textContent.trim();

                // Try to parse as numbers
                const aNum = parseFloat(aValue);
                const bNum = parseFloat(bValue);

                if (!isNaN(aNum) && !isNaN(bNum)) {
                    return direction ? aNum - bNum : bNum - aNum;
                }

                return direction ?
                    aValue.localeCompare(bValue) :
                    bValue.localeCompare(aValue);
            });

            // Re-append sorted rows
            rows.forEach(row => tbody.appendChild(row));

            // Toggle direction
            sortDirections[columnIndex] = !direction;
        }

        // Column resizing functionality (must be set up BEFORE click handlers)
        const resizers = document.querySelectorAll('.resizer');
        let currentResizer = null;
        let currentTh = null;
        let startX = 0;
        let startWidth = 0;
        let isResizing = false;
        let hasMoved = false;

        resizers.forEach(resizer => {
            resizer.addEventListener('mousedown', function(e) {
                e.stopPropagation();
                e.preventDefault();
                isResizing = true;
                hasMoved = false;
                currentResizer = resizer;
                currentTh = resizer.parentElement;
                startX = e.pageX;
                startWidth = currentTh.offsetWidth;

                document.addEventListener('mousemove', handleMouseMove);
                document.addEventListener('mouseup', handleMouseUp);
            });
        });

        // Add click handlers to table headers for sorting
        const headers = document.querySelectorAll('th[data-column]');
        headers.forEach(th => {
            th.addEventListener('click', function(e) {
                // Don't sort if we just finished resizing
                if (hasMoved) {
                    hasMoved = false;
                    return;
                }
                // Only sort if we didn't click on the resizer
                if (!e.target.classList.contains('resizer')) {
                    const columnIndex = parseInt(this.getAttribute('data-column'));
                    sortTable(columnIndex);
                }
            });
        });

        function handleMouseMove(e) {
            if (currentResizer) {
                hasMoved = true;
                const width = startWidth + (e.pageX - startX);
                if (width > 50) { // Minimum width of 50px
                    currentTh.style.width = width + 'px';
                    // Don't set minWidth - allows shrinking columns below content width
                }
            }
        }

        function handleMouseUp() {
            currentResizer = null;
            currentTh = null;
            isResizing = false;
            document.removeEventListener('mousemove', handleMouseMove);
            document.removeEventListener('mouseup', handleMouseUp);

            // Reset hasMoved after a short delay to allow click handler to check it
            setTimeout(() => {
                hasMoved = false;
            }, 10);
        }

        // Reset sort functionality
        resetSortBtn.addEventListener('click', function() {
            // Restore original row order
            originalRowOrder.forEach(row => tbody.appendChild(row));

            // Reset sort directions to initial state
            sortDirections = [true, true, true, true, true, true];
        });
    </script>
</body>
</html>
"""

    # Write HTML file with error handling
    try:
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"HTML report exported to {html_filename}\n")
        print(f"File located at {excel_directory}\n")
    except PermissionError:
        print(f"\nWARNING: Permission denied when writing to {html_filename}")
        print("Continuing without HTML report...\n")
    except OSError as e:
        print(f"\nWARNING: Failed to write HTML file: {e}")
        print("Continuing without HTML report...\n")
    except Exception as e:
        print(f"\nWARNING: Unexpected error when saving HTML file: {e}")
        print("Continuing without HTML report...\n")
else:
    print("No duplicates found to export.\n")

# Successfully completed - delete progress file
delete_progress(platform, username)

# Save filtered asset data to JSON file if enabled
if SAVE_JSON_OUTPUT:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_username = "".join(c for c in username if c.isalnum() or c in ('-', '_')).lower()
    json_filename = f"asset_data_{platform}_{safe_username}_{timestamp}.json"

    # Extract only essential fields from each asset (same as progress save)
    filtered_assets = []
    for asset in all_assets:
        essential_asset = {
            "assetId": asset.get("assetId"),
            "assetName": asset.get("assetName"),
            "dnsName": asset.get("dnsName"),
            "netbiosName": asset.get("netbiosName"),
            "macAddress": asset.get("macAddress"),
            "address": asset.get("address"),
            "inventoryListData": asset.get("inventoryListData")
        }
        filtered_assets.append(essential_asset)

    try:
        with open(json_filename, 'w') as f:
            json.dump(filtered_assets, f, indent=2)
        print(f"Asset data exported to {json_filename}\n")
    except (PermissionError, OSError) as e:
        print(f"\nWARNING: Failed to save JSON file: {e}")
        print("Continuing without JSON export...\n")

# Logout / Invalidate token by posting with token=false
logout_data = {
    "username": username,
    "password": password,
    "token": "false"
}
try:
    logout_response = requests.post(auth_url, headers=auth_headers, data=logout_data, timeout=30)
    if logout_response.status_code == 200 or logout_response.status_code == 201:
        print("Logout successful.")
    else:
        print("Logout may have failed, but token will expire in 4 hours.")
        print(logout_response.text)
except (requests.exceptions.Timeout, requests.exceptions.RequestException):
    print("Logout request failed, but token will expire in 4 hours.")
